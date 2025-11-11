import streamlit as st
import pandas as pd
from datetime import datetime
import calendar
from io import BytesIO
from openpyxl import Workbook

st.set_page_config(page_title="Nurse Roster (No-A, Custom Demand)", layout="wide")

st.title("🩺 護理師排班工具（不含 A 班／每日人力需求可自訂）")
st.caption("D=床邊出勤、O=休假；支援 ≥20 人、員工可上傳或在頁面表格輸入想休日期與每日需求")

# ========= Helpers =========
def days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]

def is_sunday(y: int, m: int, d: int) -> bool:
    return datetime(y, m, d).weekday() == 6  # Monday=0..Sunday=6

def excel_bytes(roster_df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    # header
    for j, col in enumerate(roster_df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    # rows
    for i, (_, row) in enumerate(roster_df.iterrows(), start=2):
        for j, col in enumerate(roster_df.columns, start=1):
            ws.cell(row=i, column=j, value=str(row[col]))
    # summary
    ws2 = wb.create_sheet("Summary")
    for j, col in enumerate(summary_df.columns, start=1):
        ws2.cell(row=1, column=j, value=str(col))
    for i, (_, row) in enumerate(summary_df.iterrows(), start=2):
        for j, col in enumerate(summary_df.columns, start=1):
            ws2.cell(row=i, column=j, value=row[col])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ========= Sidebar: 基本設定 =========
st.sidebar.header("排班設定")
year = st.sidebar.number_input("年份", 2024, 2100, value=2025, step=1)
month = st.sidebar.number_input("月份", 1, 12, value=11, step=1)
days = days_in_month(year, month)

# 預設用來「生成」每日需求初值（之後可在表格調整）
st.sidebar.subheader("每日需求初值（用來預填，之後可在表格改）")
default_weekday_need = st.sidebar.number_input("週一至週六 D 人數（預填）", 0, 100, 4)
default_sunday_need = st.sidebar.number_input("週日 D 人數（預填）", 0, 100, 5)

st.sidebar.subheader("限制/規則")
max_off = st.sidebar.number_input("每人每月 O（休假）上限（檢視用）", 0, 31, 8)
preserve_weekly_rest = st.sidebar.checkbox("（提示）每週至少 1 天 O", value=True)

st.sidebar.subheader("偏好處理")
strict_preferences = st.sidebar.checkbox("盡量尊重員工想休（若人力不足會被忽略）", value=True)

st.sidebar.markdown("---")
st.sidebar.subheader("資料上傳（可選）")
nurses_file = st.sidebar.file_uploader("護理師名單 CSV（欄位：id,name）", type=["csv"])
prefs_file = st.sidebar.file_uploader("想休假表單 CSV（欄位：nurse_id,date，YYYY-MM-DD）", type=["csv"])
demand_file = st.sidebar.file_uploader("每日需求 CSV（欄位：day,D_required 或 date,D_required）", type=["csv"])

# ========= 準備名單資料 =========
if nurses_file:
    nurses = pd.read_csv(nurses_file)
else:
    # 預設 20 人
    nurses = pd.DataFrame({
        "id": list(range(1, 21)),
        "name": [f"{i}號護理師" for i in range(1, 21)]
    })

# ========= 準備想休資料 =========
if prefs_file:
    prefs = pd.read_csv(prefs_file)
else:
    prefs = pd.DataFrame(columns=["nurse_id", "date"])

st.subheader("員工想休設定（本月）")
with st.expander("點此展開/編輯想休清單", expanded=False):
    month_prefix = f"{year}-{month:02d}-"
    display_prefs = prefs[prefs["date"].astype(str).str.startswith(month_prefix)].copy()
    edited = st.data_editor(
        display_prefs,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "nurse_id": st.column_config.NumberColumn("nurse_id", min_value=int(nurses["id"].min()), max_value=int(nurses["id"].max())),
            "date": st.column_config.DateColumn("date")
        },
        help="可直接在表格新增/刪除日期；系統會盡量尊重想休（若人力不足將被忽略）。"
    )
    if st.button("✅ 套用上表為本月想休"):
        other = prefs[~prefs["date"].astype(str).str.startswith(month_prefix)]
        prefs = pd.concat([other, edited], ignore_index=True)
        st.success("已套用。")

# 轉換本月想休為 map
pref_map = {int(r.id): set() for r in nurses.itertuples(index=False)}
for r in prefs.itertuples(index=False):
    try:
        dt = pd.to_datetime(r.date)
    except Exception:
        continue
    if dt.year == year and dt.month == month:
        pref_map.setdefault(int(r.nurse_id), set()).add(int(dt.day))

# ========= 準備「每日人力需求」資料表（可上傳或在表格編輯） =========
def seed_demand_df(y: int, m: int, wd_need: int, sun_need: int) -> pd.DataFrame:
    days_ = days_in_month(y, m)
    rows = []
    for d in range(1, days_ + 1):
        need = sun_need if is_sunday(y, m, d) else wd_need
        rows.append({"day": d, "date": f"{y}-{m:02d}-{d:02d}", "D_required": int(need)})
    return pd.DataFrame(rows)

if demand_file:
    raw = pd.read_csv(demand_file)
    # 兼容兩種格式：day,D_required 或 date,D_required
    if "day" in raw.columns and "D_required" in raw.columns:
        df_demand = raw.copy()
        # 若無 date 欄位，自動補
        df_demand["date"] = df_demand["day"].apply(lambda d: f"{year}-{month:02d}-{int(d):02d}")
    elif "date" in raw.columns and "D_required" in raw.columns:
        df_demand = raw.copy()
        df_demand["day"] = pd.to_datetime(df_demand["date"]).dt.day
    else:
        st.error("每日需求 CSV 欄位需為：'day,D_required' 或 'date,D_required'")
        st.stop()
else:
    df_demand = seed_demand_df(year, month, default_weekday_need, default_sunday_need)

st.subheader("每日人力需求（可直接編輯）")
with st.expander("點此展開/編輯每日需求表", expanded=True):
    df_demand = df_demand.sort_values("day").reset_index(drop=True)
    df_demand["day"] = df_demand["day"].astype(int)
    df_demand["D_required"] = df_demand["D_required"].astype(int)
    edited_demand = st.data_editor(
        df_demand,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "day": st.column_config.NumberColumn("day", min_value=1, max_value=days, step=1, help="當月第幾天"),
            "date": st.column_config.TextColumn("date", help="YYYY-MM-DD（顯示用）"),
            "D_required": st.column_config.NumberColumn("D_required", min_value=0, max_value=100, step=1, help="當天需要上 D 班的人數")
        },
        disabled=["date"],  # date 由 day 推導，避免混亂
        height=360
    )
    if st.button("✅ 套用上表為本月每日需求"):
        df_demand = edited_demand.copy()
        st.success("已套用。")

# 建立需求 map：day -> D_required
demand_map = {int(r.day): int(r.D_required) for r in df_demand.itertuples(index=False)}

# ========= 排程（D/O） =========
schedule = {int(r.id): {d: "" for d in range(1, days + 1)} for r in nurses.itertuples(index=False)}

# 先放入想休
for nid in schedule.keys():
    for d in pref_map.get(nid, set()):
        if 1 <= d <= days:
            schedule[nid][d] = "O"

assigned_D = {nid: 0 for nid in schedule.keys()}

# 逐日滿足 D 需求
for d in range(1, days + 1):
    req = max(0, int(demand_map.get(d, 0)))  # 預防空值
    # 候選：當天不是 O 的人
    candidates = [nid for nid in schedule.keys() if schedule[nid][d] != "O"]
    # 以「已分配 D 較少」優先，平均負載
    candidates.sort(key=lambda nid: (assigned_D[nid], nid))
    chosen = candidates[:req]
    for nid in chosen:
        schedule[nid][d] = "D"
        assigned_D[nid] += 1
    # 其餘標記為 O
    for nid in schedule.keys():
        if schedule[nid][d] == "":
            schedule[nid][d] = "O"

# ========= 可行性檢視 =========
def off_count(nid: int) -> int:
    return sum(1 for d in range(1, days + 1) if schedule[nid][d] == "O")

total_required_D = int(sum(max(0, int(demand_map.get(d, 0))) for d in range(1, days + 1)))
n_staff = len(schedule.keys())
avg_off = (n_staff * days - total_required_D) / n_staff if n_staff else 0

violations = [(nid, off_count(nid)) for nid in schedule.keys() if off_count(nid) > max_off]

# ========= 輸出 =========
cols = ["姓名"] + [str(d) for d in range(1, days + 1)]
id2name = {int(r.id): r.name for r in nurses.itertuples(index=False)}
rows = []
for nid in schedule.keys():
    row = {"姓名": id2name.get(nid, str(nid))}
    for d in range(1, days + 1):
        row[str(d)] = schedule[nid][d]
    rows.append(row)
roster_df = pd.DataFrame(rows).sort_values("姓名").reset_index(drop=True)

st.subheader(f"📅 {year}-{month:02d} 排班結果（D/O）")
st.dataframe(roster_df, use_container_width=True, height=520)

summary = []
for nid, name in id2name.items():
    summary.append({
        "姓名": name,
        "D天數": sum(1 for d in range(1, days + 1) if schedule[nid][d] == "D"),
        "O天數": sum(1 for d in range(1, days + 1) if schedule[nid][d] == "O"),
    })
summary_df = pd.DataFrame(summary).sort_values("姓名").reset_index(drop=True)

st.subheader("統計摘要")
st.dataframe(summary_df, use_container_width=True, height=320)

st.markdown("### 可行性檢視")
st.info(f"本月需 D 班次：**{total_required_D}**；人數：**{n_staff}**；理論平均 O/人：**{avg_off:.2f} 天**。")
if violations:
    st.warning(f"有 {len(violations)} 位人員 O 超過上限（> {max_off} 天）。在純 D/O 與目前每日需求下，要全員 ≤ {max_off} 可能數學上不可行。")
else:
    st.success("目前無 O 上限違規。")

# 下載
csv_bytes = roster_df.to_csv(index=False).encode("utf-8-sig")
st.download_button("⬇️ 下載 CSV 班表", data=csv_bytes, file_name=f"roster_{year}-{month:02d}_custom.csv", mime="text/csv")

excel_bytes_data = excel_bytes(roster_df, summary_df)
st.download_button("⬇️ 下載 Excel 班表", data=excel_bytes_data, file_name=f"roster_{year}-{month:02d}_custom.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("""
---
**使用說明**
1. 在側邊欄選擇年月，設定「預填」的週日/平日人力，系統會生成每日需求表。
2. 展開「每日人力需求」以表格編輯每一天的 D 人數（或在側邊欄上傳 CSV：`day,D_required` 或 `date,D_required`）。
3. 上傳或在頁面編輯「想休清單」（`nurse_id,date`）。系統會優先尊重，但若人力不足將忽略部分想休。
4. 產生後可下載 CSV/Excel；若顯示 O 上限違規，代表以現有人力與每日需求要把 O 壓到上限可能不可行（純數學）。
""")
